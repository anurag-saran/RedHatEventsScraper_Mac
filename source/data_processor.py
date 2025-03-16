# Data processing module for RedHat Events Scraper
import os
import logging
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from utils import ensure_directory
from config import OUTPUT_DIR

logger = logging.getLogger(__name__)

class EventDataProcessor:
    def __init__(self, output_dir=OUTPUT_DIR):
        """
        Initialize the data processor
        
        Args:
            output_dir (str): Directory for output files
        """
        self.output_dir = output_dir
        ensure_directory(output_dir)
    
    def export_to_excel(self, events, filename=None):
        """
        Export events to Excel file with improved debugging
    
        Args:
            events (list): List of event dictionaries
            filename (str): Optional filename, generated if None
        
        Returns:
            str: Path to saved file
        """
        if not events:
            logger.warning("No events to export")
            return None

        try:
            # Generate filename if not provided
            if filename is None:
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"redhat_events_all_{timestamp}.xlsx"
            
            # Check if it's a full path or just a filename
            if os.path.dirname(filename):
                # It's a full path, use it directly
                filepath = filename
                # Ensure the directory exists
                ensure_directory(os.path.dirname(filepath))
            else:
                # It's just a filename, add the output directory
                filepath = os.path.join(self.output_dir, filename)
                ensure_directory(self.output_dir)
        
            # Log the exact path where we're saving
            logger.info(f"Saving Excel file to: {filepath}")
        
            # Create a new workbook and select the active worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "RedHat Events"
        
            # Define headers - WITH New status column
            headers = [
                'New', 'Title', 'Location', 'Date Range', 'Start Date', 'End Date',
                'Link'
            ]
        
            # Write headers with styling
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
        
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
        
            # Write data with highlighting for new events
            highlight_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
            new_font = Font(bold=True, color="FF0000")  # Red, bold text

            # Write data with "New" column and highlighting
            for row_idx, event in enumerate(events, 2):
                # Check if event is new
                is_new = event.get('is_new', False)
                
                # Add "New" marker column
                new_cell = ws.cell(row=row_idx, column=1)
                new_cell.value = "NEW!" if is_new else ""
                if is_new:
                    new_cell.font = new_font
                
                # Add other columns with offset (column index +1 because of new first column)
                ws.cell(row=row_idx, column=2).value = event.get('title', 'N/A')
                ws.cell(row=row_idx, column=3).value = event.get('location', 'N/A')
                ws.cell(row=row_idx, column=4).value = event.get('date_range', 'N/A')
                ws.cell(row=row_idx, column=5).value = event.get('start_date', '')
                ws.cell(row=row_idx, column=6).value = event.get('end_date', '')
                ws.cell(row=row_idx, column=7).value = event.get('link', 'N/A')
                
                # Highlight entire row for new events
                if is_new:
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = highlight_fill
        
            # Auto-adjust column width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = min(adjusted_width, 50)
        
            # Save the workbook and log success/failure
            try:
                # If file exists, remove it first to ensure clean overwrite
                if os.path.exists(filepath):
                    try:
                        os.remove(filepath)
                        logger.info(f"Removed existing file at: {filepath}")
                    except Exception as remove_error:
                        logger.warning(f"Could not remove existing file at {filepath}: {remove_error}")
                
                wb.save(filepath)
                logger.info(f"Successfully saved Excel to: {filepath}")
                return filepath
            except Exception as save_error:
                logger.error(f"Error saving Excel file to {filepath}: {save_error}")
            
                # Try saving to current directory as fallback
                fallback_path = os.path.basename(filepath)
                logger.info(f"Attempting to save to current directory: {fallback_path}")
                wb.save(fallback_path)
                logger.info(f"Saved to fallback location: {fallback_path}")
                return fallback_path
        
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    def export_to_csv(self, events, filename=None):
        """
        Export events to CSV file
        
        Args:
            events (list): List of event dictionaries
            filename (str): Optional filename, generated if None
        
        Returns:
            str: Path to saved file
        """
        import csv
        
        if not events:
            logger.warning("No events to export to CSV")
            return None
        
        try:
            # Generate filename if not provided
            if filename is None:
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"redhat_events_{timestamp}.csv"
            
            # Create full path
            if os.path.dirname(filename):
                filepath = filename
                ensure_directory(os.path.dirname(filepath))
            else:
                filepath = os.path.join(self.output_dir, filename)
                ensure_directory(self.output_dir)
            
            logger.info(f"Saving CSV file to: {filepath}")
            
            # If file exists, remove it first to ensure clean overwrite
            if os.path.exists(filepath):
                try:
                    os.remove(filepath)
                    logger.info(f"Removed existing file at: {filepath}")
                except Exception as remove_error:
                    logger.warning(f"Could not remove existing file at {filepath}: {remove_error}")
            
            # Define headers with "New" column
            headers = [
                'New', 'Title', 'Location', 'Date Range', 'Start Date', 'End Date',
                'Link'
            ]
            
            # Write to CSV
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                
                for event in events:
                    # Check if event is new
                    is_new = event.get('is_new', False)
                    
                    writer.writerow({
                        'New': "NEW!" if is_new else "",
                        'Title': event.get('title', 'N/A'),
                        'Location': event.get('location', 'N/A'),
                        'Date Range': event.get('date_range', 'N/A'),
                        'Start Date': event.get('start_date', ''),
                        'End Date': event.get('end_date', ''),
                        'Link': event.get('link', 'N/A')
                    })
            
            logger.info(f"Successfully saved CSV to: {filepath}")
            return filepath
        
        except Exception as e:
            logger.error(f"Error creating CSV file: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    def format_for_display(self, events, max_events=10):
        """
        Format events for display in GUI
        
        Args:
            events (list): List of event dictionaries
            max_events (int): Maximum number of events to format
            
        Returns:
            str: Formatted events as string
        """
        if not events:
            return "No events found."
        
        # Limit the number of events to display
        display_events = events[:max_events]
        
        # Format as text
        formatted = []
        for i, event in enumerate(display_events, 1):
            event_text = [f"Event {i}:"]
            event_text.append(f"Title: {event.get('title', 'N/A')}")
            event_text.append(f"Type: {event.get('type', 'N/A')}")
            event_text.append(f"Date: {event.get('date_range', 'N/A')}")
            event_text.append(f"Location: {event.get('location', 'N/A')}")
            
            if event.get('description'):
                # Truncate description if too long
                desc = event.get('description')
                if len(desc) > 100:
                    desc = desc[:97] + "..."
                event_text.append(f"Description: {desc}")
            
            if event.get('link') and event.get('link') != "N/A":
                event_text.append(f"Link: {event.get('link')}")
            
            # Add "NEW!" indicator for new events
            if event.get('is_new', False):
                event_text.append("Status: NEW!")
            
            formatted.append("\n".join(event_text))
        
        # Add a note if there are more events
        if len(events) > max_events:
            formatted.append(f"\n(+ {len(events) - max_events} more events not shown)")
        
        return "\n\n".join(formatted)

    def export_to_google_sheets(self, events, spreadsheet_name='RedHat Events Tracker'):
        """
        Export events to Google Sheets
    
        Args:
            events (list): List of event dictionaries
            spreadsheet_name (str): Name of the Google Sheets file
        
        Returns:
            str: URL of the Google Sheet
        """
        try:
            from google_sheets import create_or_update_sheet
            
            # Export to Google Sheets
            sheet_url = create_or_update_sheet(events, spreadsheet_name)
            if sheet_url:
                logger.info(f"Events successfully exported to Google Sheets: {sheet_url}")
            else:
                logger.warning("Failed to export events to Google Sheets")
            
            return sheet_url
        except ImportError:
            logger.error("Google Sheets integration not available. Install required packages.")
            return None
        except Exception as e:
            logger.error(f"Error exporting to Google Sheets: {e}")
            return None